from __future__ import annotations

import argparse
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


SHEET_REF_PATTERN = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_ .-]+))!")


def stringify(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("\n", " ").strip()
    return str(value)


def iter_non_empty_rows(sheet, max_rows: int = 12, max_cols: int = 12):
    captured = 0
    for row in sheet.iter_rows():
        values = []
        for cell in row[:max_cols]:
            rendered = stringify(cell.value)
            if rendered:
                values.append(f"{cell.coordinate}={rendered}")
        if values:
            yield row[0].row, values
            captured += 1
            if captured >= max_rows:
                return


def iter_formulas(sheet, max_items: int = 25):
    captured = 0
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                yield cell.coordinate, cell.value
                captured += 1
                if captured >= max_items:
                    return


def count_formulas(sheet) -> int:
    total = 0
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                total += 1
    return total


def collect_sheet_references(sheet) -> Counter[str]:
    references: Counter[str] = Counter()
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for quoted, bare in SHEET_REF_PATTERN.findall(cell.value):
                    references[quoted or bare] += 1
    return references


def describe_sheet(sheet) -> str:
    parts = [
        f"Sheet: {sheet.title}",
        f"State: {sheet.sheet_state}",
        f"Dimensions: {sheet.calculate_dimension()}",
        f"Max row: {sheet.max_row}",
        f"Max column: {sheet.max_column}",
        f"Merged ranges: {len(sheet.merged_cells.ranges)}",
        f"Frozen panes: {sheet.freeze_panes or 'none'}",
        f"Formula cells: {count_formulas(sheet)}",
    ]

    references = collect_sheet_references(sheet)
    if references:
        formatted_refs = ", ".join(
            f"{name} ({count})" for name, count in references.most_common(8)
        )
        parts.append(f"Cross-sheet references: {formatted_refs}")

    parts.append("Top non-empty rows:")
    for row_number, values in iter_non_empty_rows(sheet):
        parts.append(f"  Row {row_number}: {' | '.join(values)}")

    formulas = list(iter_formulas(sheet))
    if formulas:
        parts.append("Formula samples:")
        for coordinate, formula in formulas:
            parts.append(f"  {coordinate}: {formula}")

    return "\n".join(parts)


def describe_workbook(workbook_path: Path) -> str:
    workbook = load_workbook(workbook_path, data_only=False)
    defined_names = list(workbook.defined_names.items())
    lines = [
        f"Workbook: {workbook_path}",
        f"Sheet count: {len(workbook.sheetnames)}",
        f"Sheets: {', '.join(workbook.sheetnames)}",
        f"Defined names: {len(defined_names)}",
    ]

    for name, defined_name in defined_names:
        attr_text = getattr(defined_name, "attr_text", None)
        if name and attr_text:
            lines.append(f"Defined name: {name} -> {attr_text}")

    for sheet in workbook.worksheets:
        lines.append("")
        lines.append("-" * 72)
        lines.append(describe_sheet(sheet))

    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect an Excel workbook in read-only mode.")
    parser.add_argument(
        "workbook",
        nargs="?",
        default="compta_2025-2026.xlsx",
        help="Path to the workbook relative to the current working directory.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).resolve()
    print(describe_workbook(workbook_path))


if __name__ == "__main__":
    main()